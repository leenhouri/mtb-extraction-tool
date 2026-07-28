"""Shared table I/O for the reporting tools (significance, interrater).

Both tools read a (human-curated) comparison table, group rows by field type, and
emit a small results table. This module holds what they share: lenient table
loading, fuzzy column lookup, the agreement formula, and rendering to the terminal
and to an .xlsx. field_type is re-exported from evaluate.py so the reporting tables
group fields exactly like the headline metrics (e.g. ``diagnosis.tumor``,
``document.type``).
"""
import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ehr.evaluation.matching import field_type  # noqa: F401  (re-exported)

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_BOLD_FONT = Font(bold=True, name="Arial", size=10)


def load_rows(path, sheet=None):
    """Read an .xlsx/.csv table into (header, list-of-dict-rows)."""
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        return header, [dict(zip(header, r)) for r in rows[1:]]
    with open(path, encoding="utf-8-sig") as f:
        rd = csv.DictReader(f)
        return rd.fieldnames, list(rd)


def find(header, *names):
    """Resolve a column by exact (case-insensitive) name, then substring fallback."""
    low = {h.strip().lower(): h for h in header if h}
    for n in names:
        if n.lower() in low:
            return low[n.lower()]
    for h in header:
        if h and any(n.lower() in h.strip().lower() for n in names):
            return h
    return None


def agreement(match, mismatch):
    """match / (match + mismatch), or 0.0 when there is nothing to score."""
    tot = match + mismatch
    return (match / tot) if tot else 0.0


def print_table(title, headers, rows):
    """Print a fixed-width text table to stdout (for quick terminal viewing)."""
    widths = [len(str(h)) for h in headers]
    for r in rows:
        for i, c in enumerate(r):
            widths[i] = max(widths[i], len(str(c)))
    print(f"\n{title}")
    print("  ".join(str(h).ljust(widths[i]) for i, h in enumerate(headers)))
    print("  ".join("-" * widths[i] for i in range(len(headers))))
    for r in rows:
        print("  ".join(str(c).ljust(widths[i]) for i, c in enumerate(r)))


def write_xlsx(path, headers, rows, sheet_name, bold_last=True):
    """Write headers + rows to a styled single-sheet .xlsx (the file deliverable)."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name
    ws.append(list(headers))
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
    for r in rows:
        ws.append(list(r))
    if bold_last and rows:
        for c in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=c).font = _BOLD_FONT
    for i, h in enumerate(headers, start=1):
        cells = [str(h)] + [str(r[i - 1]) for r in rows]
        ws.column_dimensions[get_column_letter(i)].width = max(10, max(len(c) for c in cells) + 2)
    wb.save(path)
    return path
