"""compare_xlsx.py - Render the gold-vs-LLM comparison rows as a styled .xlsx.

A single per-field 'comparison' sheet: output cells highlighted red on disagreement,
and the evaluation cell green/red with its match / mismatch / missing / extra label.
Pure presentation -- the rows are built in compare.py.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

EMPTY_LABEL = "Not documented"

HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
MISMATCH_FILL = PatternFill("solid", fgColor="FCE4E4")   # light red
MISMATCH_FONT = Font(color="C00000", name="Arial", size=10)
MATCH_FILL    = PatternFill("solid", fgColor="E6F4EA")   # light green (evaluation = match)
VAR_FONT      = Font(bold=True, name="Arial", size=10)
BASE_FONT     = Font(name="Arial", size=10)
EMPTY_FONT    = Font(italic=True, color="999999", name="Arial", size=10)
PID_FILL_A    = PatternFill("solid", fgColor="FFFFFF")
PID_FILL_B    = PatternFill("solid", fgColor="F2F2F2")
THIN          = Side(style="thin", color="D9D9D9")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_xlsx(rows, out_path, has_out1):
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "comparison"

    if has_out1:
        headers = ["patient_id", "variable", "ground_truth", "output_1", "output_2", "evaluation"]
        out1_col, out2_col, eval_col = 4, 5, 6
    else:
        headers = ["patient_id", "variable", "ground_truth", "output_2", "evaluation"]
        out1_col, out2_col, eval_col = None, 4, 5

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    zebra = {}
    for r in rows:
        if r["pid"] not in zebra:
            zebra[r["pid"]] = PID_FILL_A if len(zebra) % 2 == 0 else PID_FILL_B

    for i, r in enumerate(rows, start=2):
        if has_out1:
            vals = [r["pid"], r["var"], r["gold"], r["out1"], r["out2"], r["eval"]]
        else:
            vals = [r["pid"], r["var"], r["gold"], r["out2"], r["eval"]]
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = BORDER
            cell.font = VAR_FONT if c == 2 else BASE_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(c >= 3))
            if c in (1, 2):
                cell.fill = zebra[r["pid"]]
            if val == EMPTY_LABEL:
                cell.font = EMPTY_FONT
        # red highlights for disagreement with gold
        if has_out1 and r.get("o1_red"):
            assert out1_col is not None
            ws.cell(row=i, column=out1_col).fill = MISMATCH_FILL
            ws.cell(row=i, column=out1_col).font = MISMATCH_FONT
        if r.get("o2_red"):
            ws.cell(row=i, column=out2_col).fill = MISMATCH_FILL
            ws.cell(row=i, column=out2_col).font = MISMATCH_FONT
        # evaluation cell color
        ev_cell = ws.cell(row=i, column=eval_col)
        ev_cell.fill = MATCH_FILL if r["eval"] == "match" else MISMATCH_FILL
        if r["eval"] != "match":
            ev_cell.font = MISMATCH_FONT

    widths = [12, 34, 30, 30, 30, 12] if has_out1 else [12, 34, 30, 30, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    wb.save(out_path)
