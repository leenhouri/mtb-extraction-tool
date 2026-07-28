"""pharmacy.py - Objective check of the gold-standard therapies against pharmacy records.

The pharmacy dispensing table lists one row per dispensing event (per active ingredient).
The gold-standard medications come from a flattened comparison table (one row per field):
the medication name + start/end are read from a chosen value column (default: the
ground-truth column), so this validates the GOLD STANDARD against objective dispensing
records (rather than the raw LLM JSON).

Per patient and active ingredient, dispensing rows are aggregated into a span
(first dispensing = start, last dispensing = end). The gold medication's start/end dates
are compared against that span with a tolerance of +/- 1 month (<=2 months still
accepted). Concordance requires BOTH start and end to agree; if the gold end is absent
(ongoing therapy / "Not documented"), only the start is scored.

Each pharmacy span is capped at the patient's protocol (document) date: dispensings after
that date could not have been recorded in the gold standard, so they are excluded -- a drug
dispensed only after the protocol is dropped, and a span extending past it is truncated.

Two views are reported:
  - PRESENCE  : did the gold ingredients match the dispensed ones (precision/recall/F1)
  - DATE      : among matched ingredients, how often start / end / both agree within tolerance

Pharmacy table columns (only PAT_ID_TRUE, Datum, Wirkstoff are used; PAT_ID_NAME is PII):
  PAT_ID_NAME, PAT_ID_TRUE, Datum (DD.MM.YYYY), Wirkstoff, Dosierung, Dosis
Comparison-table columns used: patient_id, variable, and the chosen value column.
Active-ingredient canonicalisation lives in drugs.py, which also excludes supportive
medications and orally/externally dispensed agents (PARP inhibitors, oral endocrine
therapy) that the institutional intravenous pharmacy does not fill, on both sides.

Usage:
  uv run ehr-pharmacy --pharmacy apotheke.xlsx --table comparison_table.xlsx \
        --col "ground_truth (Clinician 2)" --tag run1
"""

import argparse
import csv
import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from ehr.evaluation.drugs import is_empty, canon_drug
from ehr.evaluation._tableio import load_rows, find

TOL_STRONG = 1   # months
TOL_ACCEPT = 2   # months

# Medication fields in a flattened comparison table, e.g. "doc0.treatment.med[1].startDate".
_MED_RE = re.compile(r"^(doc\d+\.treatment\.med\[[^\]]+\])\.(medicationName|startDate|endDate)$")
# Document date field, e.g. "doc0.date" (used as the protocol-date cut-off).
_DOC_DATE_RE = re.compile(r"^doc\d+\.date$")


def month_index(v):
    """Return year*12 + (month-1) for a date, or None. Accepts datetime/date objects
    (from xlsx) and strings DD.MM.YYYY, DD.MM.YY, YYYY-MM-DD, YYYY-MM, YYYY."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.year * 12 + (v.month - 1)
    if is_empty(v):
        return None
    s = str(v).strip()
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{2,4})$", s)        # DD.MM.YYYY
    if m:
        _, mo, y = int(m.group(1)), int(m.group(2)), m.group(3)
        y = int("20" + y) if len(y) == 2 else int(y)
        return y * 12 + (mo - 1)
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)                # ISO date
    if m:
        return int(m.group(1)) * 12 + (int(m.group(2)) - 1)
    m = re.match(r"^(\d{4})-(\d{2})$", s)                        # ISO month
    if m:
        return int(m.group(1)) * 12 + (int(m.group(2)) - 1)
    m = re.match(r"^(\d{4})$", s)                                # year
    if m:
        return int(m.group(1)) * 12
    return None


def pid_int(s) -> str:
    """Normalize a patient id to its integer string so '0042' == '42'."""
    m = re.search(r"\d+", str(s))
    return str(int(m.group())) if m else str(s).strip()


def concordance(a, b):
    """Return 'strong' (<=1mo), 'accept' (<=2mo), 'off' (>2mo), or None if not comparable."""
    if a is None or b is None:
        return None
    d = abs(a - b)
    if d <= TOL_STRONG:
        return "strong"
    if d <= TOL_ACCEPT:
        return "accept"
    return "off"


def _merge_drug(spans, drug, start, end):
    """Add one (start, end) for a drug into spans, keeping earliest start / latest end
    and the first available display string (drugs may recur across treatment lines)."""
    si, ei = month_index(start), month_index(end)
    sd = str(start) if not is_empty(start) else ""
    ed = str(end) if not is_empty(end) else ""
    if drug in spans:
        ps, pe, psd, ped = spans[drug]
        si = min([x for x in (si, ps) if x is not None], default=None)
        ei = max([x for x in (ei, pe) if x is not None], default=None)
        sd, ed = psd or sd, ped or ed
    spans[drug] = (si, ei, sd, ed)


# --- load pharmacy dispensing table -----------------------------------------
def _iter_table(path: str):
    """Yield dict rows from an .xlsx (openpyxl) or .csv file. Header is the first row."""
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        assert ws is not None
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip() if c is not None else "" for c in row]
                continue
            if all(c is None for c in row):
                continue
            yield {header[i]: row[i] for i in range(min(len(header), len(row)))}
        wb.close()
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            yield from csv.DictReader(f)


def load_pharmacy(path: str):
    """{pid: {drug_canon: (start_idx, end_idx, start_disp, end_disp)}}"""
    rows = defaultdict(lambda: defaultdict(list))   # pid -> drug -> [(idx, display)]
    rows_iter = list(_iter_table(path))
    fieldnames = list(rows_iter[0].keys()) if rows_iter else []
    fields = {str(k).strip().lower(): k for k in fieldnames}
    col_id   = fields.get("pat_id_true") or fields.get("pat_id") or "PAT_ID_TRUE"
    col_date = fields.get("datum") or fields.get("date") or "Datum"
    col_drug = fields.get("wirkstoff") or fields.get("drug") or "Wirkstoff"
    for r in rows_iter:
        pid = pid_int(r.get(col_id, ""))
        drug = canon_drug(r.get(col_drug, ""))
        idx = month_index(r.get(col_date, ""))
        if not drug or idx is None:
            continue
        disp = r.get(col_date, "")
        disp = disp.strftime("%d.%m.%Y") if isinstance(disp, (datetime, date)) else str(disp).strip()
        rows[pid][drug].append((idx, disp))
    out = {}
    for pid, drugs_seen in rows.items():
        out[pid] = {}
        for drug, events in drugs_seen.items():
            events.sort(key=lambda e: e[0])
            out[pid][drug] = (events[0][0], events[-1][0], events[0][1], events[-1][1])
    return out


# --- load gold medications from the comparison table ------------------------
def load_from_comparison_table(path: str, value_col=None):
    """({pid: {drug_canon: (start_idx, end_idx, start_disp, end_disp)}},
        {pid: doc_date_month_index}) read from a flattened comparison table.

    Medication name + start/end are taken from `value_col` (default: the
    ground-truth (Clinician 2) column), so the check validates the gold standard.
    The document date (doc*.date) is read from the same column and returned as a
    per-patient month index, used as the protocol-date cut-off (see _cap_pharmacy_at_doc_date)."""
    header, rows = load_rows(path)
    pid_col = find(header, "patient_id", "patient", "pid")
    var_col = find(header, "variable", "field_path", "field")
    val_col = value_col or find(header, "ground truth (clinician 2)", "ground_truth", "ground truth")
    if not (pid_col and var_col and val_col):
        raise SystemExit("Could not resolve comparison-table columns "
                         f"(patient_id={pid_col!r}, variable={var_col!r}, value={val_col!r}). "
                         f"Headers: {header}")

    slots = defaultdict(lambda: defaultdict(dict))   # pid -> med_base -> {field: value}
    doc_date = {}                                    # pid -> month index of document date
    for r in rows:
        var = r.get(var_col)
        if var is None:
            continue
        var = str(var)
        pid = pid_int(r.get(pid_col))
        m = _MED_RE.match(var)
        if m:
            slots[pid][m.group(1)][m.group(2)] = r.get(val_col)
            continue
        if _DOC_DATE_RE.match(var):
            idx = month_index(r.get(val_col))
            if idx is not None:
                # one document per patient; if several, keep the latest date
                doc_date[pid] = max(idx, doc_date[pid]) if pid in doc_date else idx

    out = {}
    for pid, bases in slots.items():
        spans = {}
        for fields in bases.values():
            drug = canon_drug(fields.get("medicationName"))
            if drug:
                _merge_drug(spans, drug, fields.get("startDate"), fields.get("endDate"))
        out[pid] = spans
    return out, doc_date


# --- compare ----------------------------------------------------------------
def _cap_pharmacy_at_doc_date(pharm, doc_date):
    """Cap each pharmacy span at the patient's document (protocol) date.

    Drops dispensing events after the document date by clamping the span end to the
    document month. If a span's START is already after the document date (the whole
    therapy was dispensed only after the protocol), the drug is removed entirely, as
    none of it was knowable at protocol time. Patients without a document date are
    left unchanged."""
    capped = {}
    for pid, spans in pharm.items():
        cut = doc_date.get(pid)
        if cut is None:
            capped[pid] = spans
            continue
        new_spans = {}
        for drug, (si, ei, sd, ed) in spans.items():
            if si is not None and si > cut:
                continue                      # entire therapy dispensed after the protocol
            new_ei = ei
            if ei is not None and ei > cut:
                new_ei = cut                  # clamp end to the document month
            new_spans[drug] = (si, new_ei, sd, ed)
        capped[pid] = new_spans
    return capped


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pharmacy", required=True, help="Pharmacy dispensing table (.xlsx or .csv)")
    ap.add_argument("--table", required=True, help="Comparison table (.xlsx/.csv) with the gold values")
    ap.add_argument("--col", default=None,
                    help="Value column in the table to read the gold medications from "
                         "(default: the ground-truth (Clinician 2) column)")
    ap.add_argument("--tag", required=True)
    args = ap.parse_args()

    pharm = load_pharmacy(args.pharmacy)
    gold, doc_date = load_from_comparison_table(args.table, args.col)

    # Cap each pharmacy span at the patient's protocol (document) date: dispensings after the
    # protocol could not have been recorded in the gold standard, so they are excluded.
    pharm = _cap_pharmacy_at_doc_date(pharm, doc_date)

    rows = []
    # presence counts (micro, over patients present in BOTH sources)
    tp = fp = fn = 0
    start_counts = defaultdict(int)   # strong/accept/off
    end_counts = defaultdict(int)
    both_ok = both_total = 0

    patients = sorted(set(pharm) & set(gold), key=lambda x: int(x) if x.isdigit() else 0)
    for pid in patients:
        d_gold, d_ph = gold.get(pid, {}), pharm.get(pid, {})
        for drug in sorted(set(d_gold) | set(d_ph)):
            in_g, in_p = drug in d_gold, drug in d_ph
            if in_g and in_p:
                tp += 1
                gs, ge, gsd, ged = d_gold[drug]
                ps, pe, psd, ped = d_ph[drug]
                sc = concordance(gs, ps)
                ec = concordance(ge, pe)
                if sc: start_counts[sc] += 1
                if ec:
                    end_counts[ec] += 1
                    both_total += 1
                    if sc in ("strong", "accept") and ec in ("strong", "accept"):
                        both_ok += 1
                status = "matched"
            elif in_g:
                fp += 1; gs, ge, gsd, ged = d_gold[drug]; psd = ped = ""; sc = ec = None
                status = "gold_only"
            else:
                fn += 1; ps, pe, psd, ped = d_ph[drug]; gsd = ged = ""; sc = ec = None
                status = "dispensed_only"
            rows.append({
                "patient": pid, "drug": drug, "status": status,
                "gold_start": gsd, "dispensed_start": psd, "start": sc or "",
                "gold_end": ged, "dispensed_end": ped, "end": ec or "",
            })

    prec = tp / (tp + fp) if (tp + fp) else 0.0
    rec  = tp / (tp + fn) if (tp + fn) else 0.0
    f1   = 2 * prec * rec / (prec + rec) if (prec + rec) else 0.0

    out_csv = f"pharmacy_concordance_{args.tag}.csv"
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["patient","drug","status","gold_start","dispensed_start","start","gold_end","dispensed_end","end"])
        w.writeheader(); w.writerows(rows)

    n_start = sum(start_counts.values())
    start_ok = start_counts["strong"] + start_counts["accept"]
    end_ok = end_counts["strong"] + end_counts["accept"]

    summary = {
        "patients_compared": len(patients),
        "presence": {"tp": tp, "fp": fp, "fn": fn, "precision": round(prec,3), "recall": round(rec,3), "f1": round(f1,3)},
        "start_date": {"matched": n_start, "within_1mo": start_counts["strong"], "within_2mo": start_ok, "off": start_counts["off"],
                       "rate": round(start_ok / n_start, 3) if n_start else 0.0},
        "end_date": {"comparable": both_total, "within_2mo": end_ok, "off": end_counts["off"],
                     "rate": round(end_ok / both_total, 3) if both_total else 0.0},
        "both_dates": {"comparable": both_total, "ok": both_ok, "rate": round(both_ok / both_total, 3) if both_total else 0.0},
    }
    Path(f"pharmacy_concordance_{args.tag}.json").write_text(json.dumps(summary, indent=2))

    print(f"Patients compared (in both sources): {len(patients)}")
    print("\nPRESENCE (gold vs dispensed ingredients):")
    print(f"  TP={tp}  FP={fp}  FN={fn}  |  precision={prec:.2f}  recall={rec:.2f}  F1={f1:.2f}")
    print(f"\nDATE CONCORDANCE (among {tp} matched ingredients, tolerance <=1mo / <=2mo):")
    print(f"  start: {start_ok}/{n_start} within 2mo ({summary['start_date']['rate']:.2f})  [{start_counts['strong']} within 1mo]")
    print(f"  end:   {end_ok}/{both_total} within 2mo ({summary['end_date']['rate']:.2f})  ({n_start-both_total} ends not comparable)")
    print(f"  both:  {both_ok}/{both_total} ({summary['both_dates']['rate']:.2f})")
    print(f"\nWritten: {out_csv}, pharmacy_concordance_{args.tag}.json")
    print("Note: FP (gold_only) may include orally/externally administered agents absent from this table.")


if __name__ == "__main__":
    main()